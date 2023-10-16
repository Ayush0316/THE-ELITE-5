from django.shortcuts import render
from .models import H1BDisclosure
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from django.http import HttpResponse
import psycopg2
import csv
# from django import forms
# from django.shortcuts import render

# class MyForm(forms.Form):
#     file = forms.FileField()

# def submit_form(request):
#     if request.method == 'POST':
#         form = MyForm(request.POST, request.FILES)
#         if form.is_valid():
#             uploaded_file = form.cleaned_data['file']

#             # Call your Python function here, passing the uploaded file.
#             # For example, your_function(uploaded_file)
#     else:
#         form = MyForm()
#     return render(request, 'your_template.html', {'form': MyForm()})



def saving_to_database(sheet):
    # Database connection parameters
    db_params = {
        'host': 'localhost',
        'port': 5432,
        'database': 'STGI',
        'user': 'postgres',
        'password': 'AYUSH28@'
    }


    # # Table name in the database
    # table_name = 'NEXT_TESTING'
    table_name = 'NEXT_TESTING'

    # Establish a database connection
    conn = psycopg2.connect(**db_params)
    cursor = conn.cursor()

    # Create a table to store the CSV data (if it doesn't exist)
    create_table_query = """
        CREATE TABLE IF NOT EXISTS {} (
        CASE_NUMBER VARCHAR PRIMARY KEY,
        CASE_STATUS VARCHAR,
        CASE_SUBMITTED DATE,
        DECISION_DATE DATE,
        VISA_CLASS VARCHAR,
        EMPLOYMENT_START_DATE DATE,
        EMPLOYMENT_END_DATE DATE,
        EMPLOYER_NAME VARCHAR,
        EMPLOYER_BUSINESS_DBA VARCHAR,
        EMPLOYER_ADDRESS VARCHAR,
        EMPLOYER_CITY VARCHAR,
        EMPLOYER_STATE VARCHAR,
        EMPLOYER_POSTAL_CODE VARCHAR,
        EMPLOYER_COUNTRY VARCHAR,
        EMPLOYER_PROVINCE VARCHAR,
        EMPLOYER_PHONE VARCHAR,
        EMPLOYER_PHONE_EXT VARCHAR,
        AGENT_REPRESENTING_EMPLOYER VARCHAR,
        AGENT_ATTORNEY_NAME VARCHAR,
        AGENT_ATTORNEY_CITY VARCHAR,
        AGENT_ATTORNEY_STATE VARCHAR,
        JOB_TITLE VARCHAR,
        SOC_CODE VARCHAR,
        SOC_NAME VARCHAR,
        NAICS_CODE VARCHAR,
        TOTAL_WORKERS INT,
        NEW_EMPLOYMENT INT,
        CONTINUED_EMPLOYMENT INT,
        CHANGE_PREVIOUS_EMPLOYMENT INT,
        NEW_CONCURRENT_EMPLOYMENT INT,
        CHANGE_EMPLOYER INT,
        AMENDED_PETITION INT,
        FULL_TIME_POSITION VARCHAR,
        PREVAILING_WAGE FLOAT,
        PW_UNIT_OF_PAY VARCHAR,
        PW_WAGE_LEVEL VARCHAR,
        PW_SOURCE VARCHAR,
        PW_SOURCE_YEAR INT,
        PW_SOURCE_OTHER VARCHAR,
        WAGE_RATE_OF_PAY_FROM FLOAT,
        WAGE_RATE_OF_PAY_TO FLOAT,
        WAGE_UNIT_OF_PAY VARCHAR,
        H1B_DEPENDENT VARCHAR,
        WILLFUL_VIOLATOR VARCHAR,
        SUPPORT_H1B VARCHAR,
        LABOR_CON_AGREE VARCHAR,
        PUBLIC_DISCLOSURE_LOCATION VARCHAR,
        WORKSITE_CITY VARCHAR,
        WORKSITE_COUNTY VARCHAR,
        WORKSITE_STATE VARCHAR,
        WORKSITE_POSTAL_CODE VARCHAR,
        ORIGINAL_CERT_DATE DATE
        );
    """.format(table_name)
    cursor.execute(create_table_query)

    cursor.execute(f"DELETE FROM {table_name}")

    salary_cols = [33,39,40]

    print("starting reading")
    i = 0
    for row in sheet.iter_rows(values_only=True):
        if(i == 0):
            i += 1
            continue
        print(i)
        i += 1
        # for data in salary_cols:
        #     pre = row[data].replace(",","")
        #     print(pre)
        #     if (pre != ""):
        #         row[data] = float(pre)

        # print(i)
        # print(row)
        cleaned_row = [value if value != "" else None for value in row]
        # print(cleaned_row)
        insert_query = f"INSERT INTO {table_name} VALUES ({', '.join(['%s']*len(cleaned_row))});"
        cursor.execute(insert_query, cleaned_row)
    print("done")

    # Commit the changes and close the cursor and connection
    conn.commit()
    cursor.close()
    conn.close()

obj = None
def landing(request):
        obj = H1BDisclosure.objects.all()
        return render(request,"landing_page/landing.html") 


options = {
    "Option 1": "WAGE_RATE_OF_PAY_FROM",
    "Option 2": "WAGE_RATE_OF_PAY_TO",
    "Option 3": "PREVAILING_WAGE"
}


def mean(input_list):
    non_none_values = [value for value in input_list if value is not None]
    
    if len(non_none_values) == 0:
        return None

    total = sum(non_none_values)
    mean = total / len(non_none_values)
    return mean


def median(input_list):
    input_list = [value for value in input_list if value is not None]
    sorted_list = sorted(input_list)
    n = len(sorted_list)

    if n % 2 == 0:
        middle1 = sorted_list[n // 2 - 1]
        middle2 = sorted_list[n // 2]
        median = (middle1 + middle2) / 2
    else:
        median = sorted_list[n // 2]
    return median


def pre25(input_list):
    input_list = [value for value in input_list if value is not None]
    sorted_list = sorted(input_list)
    n = len(sorted_list)
    index = int(0.25 * (n - 1))

    if n % 4 == 0:
        value1 = sorted_list[index]
        value2 = sorted_list[index + 1]
        percentile_25 = (value1 + value2) / 2
    else:
        percentile_25 = sorted_list[index]

    return percentile_25

def pre75(input_list):
    input_list = [value for value in input_list if value is not None]
    sorted_list = sorted(input_list)
    n = len(sorted_list)
    index = int(0.75 * (n - 1))

    if n % 4 == 0:
        value1 = sorted_list[index]
        value2 = sorted_list[index + 1]
        percentile_75 = (value1 + value2) / 2
    else:
        percentile_75 = sorted_list[index]

    return percentile_75

def working_on_data(col_name, func):
    data = H1BDisclosure.objects.values_list(col_name.lower(), flat=True)
    data_list = list(data)
    ans = 0
    if (func == "mean"):
        ans = mean(data_list)
    elif (func == "median"):
        ans = median(data_list)
    elif (func == "pre25"):
        ans = pre25(data_list)
    else:
        ans = pre75(data_list)
    return ans
    
def count_non_null_values(row):
    count = 0
    for value in row:
        if value is not None:
            count += 1
    return count


def total(ans):
    data = H1BDisclosure.objects.values_list(ans.lower(), flat=True)
    data_list = list(data)
    ans = count_non_null_values(data_list)
    return ans

def data(request):
    if request.method == "POST":
        # print(request.header)
        if (len(request.POST.keys())) >= 2:
            key = list(request.POST.keys())
            ans = request.POST.get(key[-1])
            column_name = options.get(ans)

            answer = 0
            if(column_name != None):
                answer = working_on_data(column_name, key)
            else:
                answer = total(ans)
            print(answer)
        else:
            uploaded_file = request.FILES['document']

            
            if uploaded_file.name.endswith('.xlsx'):

                # Load the Excel file using openpyxl
                wb = load_workbook(uploaded_file, read_only=True)

                # Assume you want to convert the first sheet to CSV
                sheet = wb.active
                saving_to_database(sheet)
                obj = H1BDisclosure.objects.all()
   
        return render(request,"data_analysis/homePage.html") 

    return render(request,"data_analysis/homePage.html") 

